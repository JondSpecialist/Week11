import openpyxl as op
import urllib
import os.path
from os import listdir
import concurrent.futures
import time


excel_file = 'GRI_2017_2020.xlsx'
folder = 'pdf_download/'

def get_rows_excel(filename):
    '''
    creates a generator that loads (the first sheet of) an excel file row by row.
    Utilises openpyxl.
    '''
    workbook = op.load_workbook(filename)
    worksheet = workbook[workbook.sheetnames[0]]
    yield from worksheet.iter_rows()

def save_pdf_url(name, url, path):
    '''
    Downloads and saves a pdf from a given url.
    Utilises the urllib module.
    '''
    req = urllib.request.Request(url)
    req.add_header('Accept', 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',)
    req.add_header('Accept-Encoding', 'gzip, deflate')
    req.add_header('Accept-Language', 'en-US,en;q=0.9')
    req.add_header('User-Agent', 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/103.0.0.0 Safari/537.36')
    response = urllib.request.urlopen(req,timeout=5)
    pdf = response.read()
    with open(path + name + '.pdf', 'wb') as file:
        file.write(pdf)


def metadata_excel():
    """
    Creates the excel file used to store information about the download process.
    If the file already exists it is simply opened and loaded.
    Uses openpyxl and builtin functions.
    """
    if os.path.isfile('MetaData.xlsx') == True:
        wb = op.load_workbook('MetaData.xlsx')
    else:
        wb = op.workbook.Workbook()
    sheet = wb[wb.sheetnames[0]]
    sheet.cell(1, 1,'BRnum')
    sheet.cell(1, 2,'Downloaded')
    sheet.cell(1, 4,'Primary URL')
    sheet.cell(1, 5,'Secondary URL')
    return wb


##### Old single thread downloader, the multi-thread variant in use can be found below
def GRI_pdf_downloader(file, path):
    """
    Main function downloading GRI pdf files.
    A yield function is used to minimize memory usage.
    Pdf files already in download folder are skipped.
    Updates the MetaData.xlsx file with info about download succes and urls.
    """
    generator = get_rows_excel(file)
    meta = metadata_excel()
    meta_sheet = meta[meta.sheetnames[0]]
    next(generator) # skip first row containing headers
    wb = op.load_workbook(file)
    max_rows = wb[wb.sheetnames[0]].max_row - 1 # -1 for header row
    for i in range(max_rows):
        
        if i != 0:
            print('- - - - - - -')
        # Generate next row:
        row = next(generator)
        # Copy BR# and urls from GRI to MetaData.xlsx:
        meta_sheet.cell(i+2, 1, row[0].value)
        meta_sheet.cell(i+2, 4, row[37].value)
        meta_sheet.cell(i+2, 5, row[38].value)
        
        defec = False
        
        # Check if pdf already downloaded or links marked as defective:
        if str(row[0].value) + '.pdf' in listdir(path):
            print(row[0].value+': PDF already downloaded')
        elif meta_sheet.cell(i+2,2).value == 'defective':
            print(row[0].value+': Download links marked as defective')
        # if meta_sheet.cell(i+2,2).value == 'yes' or meta_sheet.cell(i+2,2).value == 'defective':
        #     if meta_sheet.cell(i+2,2).value == 'yes':
        #         print(row[0].value+': PDF already downloaded')
        #     else:
        #         print(row[0].value+': Download links marked as defective')
        
        # try downloading from urls:
        else:
            try:
                name = row[0].value
                url = row[37].value
                save_pdf_url(name, url, path)
                print(row[0].value+': Succesfully downloaded')
                meta_sheet.cell(i+2, 2,'yes')
            except Exception as e1:
                print('Error!')
                print(e1)
                for s in ['404','403','certificate verify failed']:
                    if s in str(e1):
                        defec = True
                try:
                    name = row[0].value
                    url = row[38].value
                    save_pdf_url(name, url, path)
                    print(row[0].value+': Succesfully downloaded')
                    meta_sheet.cell(i+2, 2,'yes')
                except Exception as e2:
                    for s in ['404','403','certificate verify failed','The system cannot find the path specified']:
                        if s in str(e2):
                            defec = True
                    print('Error!')
                    print(e2)
                    print(row[0].value+': Not downloaded' )
                    if defec == True:
                        meta_sheet.cell(i+2, 2,'defective')
                    else:
                        meta_sheet.cell(i+2, 2,'no')
    meta.save('MetaData.xlsx')

#GRI_pdf_downloader(excel_file, folder)



def GRI_pdf_multi_downloader(file, path):
    """
    Main function downloading GRI pdf files. NOW WITH MULTITHREADING!
    A yield function is used to minimize memory usage.
    Pdf files already in download folder are skipped.
    Updates the MetaData.xlsx file with info about download succes and urls.
    """
    start_time = time.time()
    
    generator = get_rows_excel(file)
    meta = metadata_excel()
    meta_sheet = meta[meta.sheetnames[0]]
    next(generator) # skip first row containing headers
    wb = op.load_workbook(file)
    max_rows = wb[wb.sheetnames[0]].max_row 

    
    def download_row(index):
        """
        Downloads PDF's using try-blocks to attempt with 2 urls provided in GRI_2017_2020.xlsx.
        Updates the MetaData.xlsx file with the appropriate data.
        This function is used by the concurrent.futures module to run several downloads simultaniously.
        """
        # Generate next row:
        row = next(generator)
        
        # Copy BR# and urls from GRI to MetaData.xlsx:
        meta_sheet.cell(index, 1, row[0].value)
        meta_sheet.cell(index, 4, row[37].value)
        meta_sheet.cell(index, 5, row[38].value)
        
        defec = False
        
        # Check if pdf already downloaded or links marked as defective:
        if str(row[0].value) + '.pdf' in listdir(path):
            print(row[0].value+': PDF already downloaded\n- - - - - - -')
            meta_sheet.cell(index, 2,'yes')
        elif meta_sheet.cell(index,2).value == 'defective':
            print(row[0].value+': Download links marked as defective\n- - - - - - -')      
        else:
            try:
                name = row[0].value
                url = row[37].value
                save_pdf_url(name, url, path)
                print(row[0].value+': Succesfully downloaded\n- - - - - - -')
                meta_sheet.cell(index, 2,'yes')
            except Exception as e1:
                for s in ['404','403','certificate verify failed']:
                    if s in str(e1):
                        defec = True
                try:
                    name = row[0].value
                    url = row[38].value
                    save_pdf_url(name, url, path)
                    print('Error! '+ str(e1) +'\n' + row[0].value+': Succesfully downloaded\n- - - - - - -')
                    meta_sheet.cell(index, 2,'yes')
                except Exception as e2:
                    for s in ['404','403','certificate verify failed','The system cannot find the path specified']:
                        if s in str(e2):
                            defec = True

                    print('Error! '+ str(e1) +'\n' + 'Error! '+ str(e2) +'\n' + row[0].value+': Not downloaded\n- - - - - - -' )
                    if defec == True:
                        meta_sheet.cell(index, 2,'defective')
                    else:
                        meta_sheet.cell(index, 2,'no')    
        
    
    # run the download_row function with multithreading usng concurent.futures:
    if __name__ == '__main__':
        index_range = list(range(2, max_rows+1)) 
        
        with concurrent.futures.ThreadPoolExecutor(max_workers=16) as executor:
            executor.map(download_row, index_range)
    
    meta.save('MetaData.xlsx')
    
    end_time =time.time()
    duration = end_time - start_time
    print('Runtime: '+str(duration)+' seconds')

     
GRI_pdf_multi_downloader(excel_file, folder)
    
